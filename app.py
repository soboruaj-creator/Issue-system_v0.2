from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime, timedelta
import sqlite3
import pandas as pd
import re
import os
import io
import tempfile
from werkzeug.utils import secure_filename
import json

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
app.config['SECRET_KEY'] = 'voc-management-secret-key'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 3600

# 업로드 폴더 생성
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def read_excel_with_drm(file_storage):
    """DRM 우회 엑셀 읽기 함수"""
    temp_file_path = None
    last_error = None
    
    # 방법 1: 메모리 스트림 (openpyxl)
    try:
        file_storage.seek(0)
        file_bytes = io.BytesIO(file_storage.read())
        df = pd.read_excel(file_bytes, engine='openpyxl')
        print("DRM 처리 성공: 메모리 스트림 (openpyxl)")
        return df
    except Exception as e:
        last_error = f"openpyxl 메모리 스트림 실패: {str(e)}"
        print(f"방법 1 실패: {last_error}")
    
    # 방법 2: xlrd 엔진 (메모리 스트림)
    try:
        file_storage.seek(0)
        file_bytes = io.BytesIO(file_storage.read())
        df = pd.read_excel(file_bytes, engine='xlrd')
        print("DRM 처리 성공: 메모리 스트림 (xlrd)")
        return df
    except Exception as e:
        last_error = f"xlrd 메모리 스트림 실패: {str(e)}"
        print(f"방법 2 실패: {last_error}")
    
    # 방법 3: pyxlsb 엔진 (메모리 스트림)
    try:
        file_storage.seek(0)
        file_bytes = io.BytesIO(file_storage.read())
        df = pd.read_excel(file_bytes, engine='pyxlsb')
        print("DRM 처리 성공: 메모리 스트림 (pyxlsb)")
        return df
    except Exception as e:
        last_error = f"pyxlsb 메모리 스트림 실패: {str(e)}"
        print(f"방법 3 실패: {last_error}")
    
    # 방법 4: 임시 파일 저장 (openpyxl)
    temp_file_path = None
    try:
        file_storage.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file_path = temp_file.name
            file_storage.save(temp_file)
        
        df = pd.read_excel(temp_file_path, engine='openpyxl')
        print("DRM 처리 성공: 임시 파일 (openpyxl)")
        return df
    except Exception as e:
        last_error = f"임시 파일 openpyxl 실패: {str(e)}"
        print(f"방법 4 실패: {last_error}")
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except:
                pass
    
    # 방법 5: 임시 파일 저장 (xlrd)
    temp_file_path = None
    try:
        file_storage.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as temp_file:
            temp_file_path = temp_file.name
            file_storage.save(temp_file)
        
        df = pd.read_excel(temp_file_path, engine='xlrd')
        print("DRM 처리 성공: 임시 파일 (xlrd)")
        return df
    except Exception as e:
        last_error = f"임시 파일 xlrd 실패: {str(e)}"
        print(f"방법 5 실패: {last_error}")
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except:
                pass
    
    # 방법 6: 임시 파일 저장 (pyxlsb)
    temp_file_path = None
    try:
        file_storage.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsb') as temp_file:
            temp_file_path = temp_file.name
            file_storage.save(temp_file)
        
        df = pd.read_excel(temp_file_path, engine='pyxlsb')
        print("DRM 처리 성공: 임시 파일 (pyxlsb)")
        return df
    except Exception as e:
        last_error = f"임시 파일 pyxlsb 실패: {str(e)}"
        print(f"방법 6 실패: {last_error}")
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except:
                pass
    
    # 방법 7: 기본 엔진 (메모리 스트림)
    try:
        file_storage.seek(0)
        file_bytes = io.BytesIO(file_storage.read())
        df = pd.read_excel(file_bytes)
        print("DRM 처리 성공: 기본 엔진 (메모리 스트림)")
        return df
    except Exception as e:
        last_error = f"기본 엔진 메모리 스트림 실패: {str(e)}"
        print(f"방법 7 실패: {last_error}")
    
    # 방법 8: 기본 엔진 (임시 파일)
    temp_file_path = None
    try:
        file_storage.seek(0)
        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
            temp_file_path = temp_file.name
            file_storage.save(temp_file)
        
        df = pd.read_excel(temp_file_path)
        print("DRM 처리 성공: 기본 엔진 (임시 파일)")
        return df
    except Exception as e:
        last_error = f"기본 엔진 임시 파일 실패: {str(e)}"
        print(f"방법 8 실패: {last_error}")
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except:
                pass
    
    # 모든 방법 실패
    error_msg = f"""
    모든 DRM 처리 방법이 실패했습니다.
    
    해결 방법:
    1. 엑셀 파일을 열고 '다른 이름으로 저장' 선택
    2. 저장 시 'Excel 통합 문서(*.xlsx)' 형식으로 저장
    3. 저장된 파일로 다시 업로드 시도
    
    마지막 오류: {last_error}
    """
    raise Exception(error_msg)

def init_db():
    """데이터베이스 초기화"""
    conn = sqlite3.connect('voc_data.db')
    c = conn.cursor()
    
    # 사내 VOC 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS internal_voc (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        case_code TEXT UNIQUE NOT NULL,
        title TEXT,
        model_name TEXT,
        model_no TEXT,
        chipset TEXT,
        build_version TEXT,
        os_version TEXT,
        issue_type TEXT,
        problem TEXT,
        original_content TEXT,
        reproduction_path TEXT,
        resolver TEXT,
        resolve_option TEXT,
        cause TEXT,
        solution TEXT,
        third_party_app TEXT,
        created_date TEXT,
        uploaded_date TEXT
    )''')
    
    # 칩셋 매핑 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS chipset_mapping (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        model_name TEXT UNIQUE NOT NULL,
        chipset TEXT NOT NULL
    )''')
    
    # 3rd party 앱 키워드 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS app_keywords (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        app_name TEXT NOT NULL,
        keywords TEXT NOT NULL
    )''')
    
    # 댓글 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS comments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        voc_id INTEGER NOT NULL,
        voc_type TEXT NOT NULL,
        comment TEXT NOT NULL,
        created_date TEXT NOT NULL,
        FOREIGN KEY (voc_id) REFERENCES internal_voc(id)
    )''')
    
    # 알림 설정 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS notification_settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        enabled INTEGER DEFAULT 1,
        notification_time TEXT DEFAULT '09:00'
    )''')
    
    # 월별 메모 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS monthly_memos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        month TEXT UNIQUE NOT NULL,
        memo TEXT NOT NULL,
        created_date TEXT NOT NULL,
        updated_date TEXT NOT NULL
    )''')
    
    # 주별 메모 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS weekly_memos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        week TEXT UNIQUE NOT NULL,
        memo TEXT NOT NULL,
        created_date TEXT NOT NULL,
        updated_date TEXT NOT NULL
    )''')
    
    # 모델별 월별 메모 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS model_monthly_memos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        model_name TEXT NOT NULL,
        month TEXT NOT NULL,
        memo TEXT NOT NULL,
        created_date TEXT NOT NULL,
        updated_date TEXT NOT NULL,
        UNIQUE(model_name, month)
    )''')
    
    # Q-data 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS q_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_date TEXT,
        process_type TEXT,
        repair_name TEXT,
        repair_detail TEXT,
        detail_content TEXT,
        model_name TEXT,
        serial_number TEXT,
        log_id TEXT,
        sw_before TEXT,
        sw_after TEXT,
        uploaded_date TEXT,
        UNIQUE(serial_number, log_id)
    )''')
    
    conn.commit()
    conn.close()

def extract_watch_model(title):
    """H열에서 '워치' 단어 추출"""
    if not title:
        return None
    
    # 'watch' + 숫자 패턴 찾기 (대소문자 구분 없음)
    match = re.search(r'watch\d*', title, re.IGNORECASE)
    if match:
        # 대문자로 변환
        return match.group(0).upper()
    
    # '워치' + 숫자 패턴 찾기 (대소문자 구분 없음)
    match = re.search(r'워치\d*', title, re.IGNORECASE)
    if match:
        # 대문자로 변환
        return match.group(0).upper()
    
    # 'watch' 단어만 있는 경우
    if 'watch' in title.lower():
        return 'WATCH'
    
    # '워치' 단어만 있는 경우
    if '워치' in title.lower():
        return '워치'
    
    return None

def map_model_name(model_name):
    """모델명 매핑"""
    if not model_name:
        return None
    
    # 대소문자 구분 없이 매핑
    model_name_upper = model_name.upper()
    
    # 워치울
    if model_name_upper in ['SM-L705N', 'SM-L705']:
        return '워치울'
    
    # 워치7
    if model_name_upper in ['SM-L310N', 'SM-L310', 'SM-L305N', 'WATCH7']:
        return '워치7'
    
    # 워치4
    if model_name_upper in ['SM-R890', 'SM-R870']:
        return '워치4'
    
    # 워치6
    if model_name_upper in ['SM-R935N', 'SM-R960', 'SM-R950', 'SM-R940', 'WATCH6']:
        return '워치6'
    
    return model_name

def extract_model_from_title(title):
    """타이틀에서 모델명 추출 (SM-XXX 형식)"""
    if not title:
        return None
    match = re.search(r'SM-[A-Z0-9]{4,5}', title, re.IGNORECASE)
    return match.group(0).upper() if match else None

def extract_model_from_reproduction(text):
    """재현경로에서 [Model No.] 추출"""
    if not text:
        return None
    match = re.search(r'\[Model No\.\]\s*([^\[\n]+)', text, re.IGNORECASE)
    if match:
        model = match.group(1).strip()
        # SM- 형식 찾기
        sm_match = re.search(r'SM-[A-Z0-9]{4,5}', model, re.IGNORECASE)
        return sm_match.group(0).upper() if sm_match else model
    return None

def extract_build_version(text):
    """재현경로에서 [Build No.] 뒤 3글자 추출"""
    if not text:
        return None
    match = re.search(r'\[Build No\.\]\s*([^\[\n]+)', text, re.IGNORECASE)
    if match:
        build = match.group(1).strip()
        # 마지막 3글자 추출
        return build[-3:] if len(build) >= 3 else build
    return None

def extract_os_version(text):
    """재현경로에서 [OS Ver.] 추출"""
    if not text:
        return None
    match = re.search(r'\[OS Ver\.\]\s*([^\[\n]+)', text, re.IGNORECASE)
    return match.group(1).strip() if match else None

def extract_original_content(text):
    """재현경로에서 [Original Contents] 추출"""
    if not text:
        return None
    match = re.search(r'\[Original Contents\]\s*([^\[]+)', text, re.IGNORECASE)
    return match.group(1).strip() if match else None

def detect_issue_type(problem, reproduction):
    """외부/내부 이슈 구분"""
    text = f"{problem or ''} {reproduction or ''}"
    external_keywords = ['samsung members', 'k zone', 'rdm']
    
    for keyword in external_keywords:
        if keyword.lower() in text.lower():
            return '외부이슈'
    return '내부이슈'

def detect_third_party_app(text):
    """텍스트에서 3rd party 앱 감지"""
    if not text:
        return None
    
    conn = sqlite3.connect('voc_data.db')
    c = conn.cursor()
    c.execute("SELECT app_name, keywords FROM app_keywords")
    apps = c.fetchall()
    conn.close()
    
    detected_apps = []
    for app_name, keywords_str in apps:
        keywords = [k.strip() for k in keywords_str.split(',')]
        for keyword in keywords:
            if keyword.lower() in text.lower():
                detected_apps.append(app_name)
                break
    
    return ', '.join(detected_apps) if detected_apps else None

def get_chipset_for_model(model_name):
    """모델명으로 칩셋 조회"""
    if not model_name:
        return None
    
    conn = sqlite3.connect('voc_data.db')
    c = conn.cursor()
    c.execute("SELECT chipset FROM chipset_mapping WHERE model_name = ?", (model_name,))
    result = c.fetchone()
    conn.close()
    
    return result[0] if result else None

def normalize_chipset_name(chipset):
    """칩셋명 정규화 (유사도 비교용)"""
    if not chipset:
        return ""
    
    # JDM T618은 예외 처리
    if chipset.strip() == "JDM T618":
        return "jdm t618"
    
    # SM으로 시작하는 칩셋은 예외 처리
    if chipset.strip().upper().startswith("SM"):
        return chipset.strip().lower()
    
    # 소문자로 변환하고 공백, 괄호, 특수문자 제거
    normalized = chipset.lower()
    normalized = re.sub(r'[\s\(\)\[\]]+', '', normalized)
    normalized = re.sub(r'[^a-z0-9]', '', normalized)
    
    return normalized

def find_similar_chipset(new_chipset, existing_chipsets, threshold=0.7):
    """유사한 칩셋명 찾기"""
    if not new_chipset or not existing_chipsets:
        return None
    
    new_normalized = normalize_chipset_name(new_chipset)
    
    for existing_chipset in existing_chipsets:
        existing_normalized = normalize_chipset_name(existing_chipset)
        
        # 정확히 일치
        if new_normalized == existing_normalized:
            return existing_chipset
        
        # 부분 문자열 일치
        if new_normalized in existing_normalized or existing_normalized in new_normalized:
            return existing_chipset
        
        # 유사도 계산 (간단한 문자열 유사도)
        similarity = calculate_string_similarity(new_normalized, existing_normalized)
        if similarity >= threshold:
            return existing_chipset
    
    return None

def calculate_string_similarity(str1, str2):
    """두 문자열 간의 유사도 계산 (0-1)"""
    if not str1 or not str2:
        return 0.0
    
    # 더 긴 문자열을 기준으로 유사도 계산
    longer = str1 if len(str1) >= len(str2) else str2
    shorter = str2 if len(str1) >= len(str2) else str1
    
    if len(longer) == 0:
        return 1.0
    
    # 공통 문자 수 계산
    common_chars = set(longer) & set(shorter)
    similarity = len(common_chars) / len(set(longer))
    
    return similarity

def select_longer_chipset(chipset1, chipset2):
    """두 칩셋명 중 더 긴 것을 선택"""
    if not chipset1:
        return chipset2
    if not chipset2:
        return chipset1
    
    return chipset1 if len(chipset1) >= len(chipset2) else chipset2

def merge_similar_chipsets(chipsets):
    """유사한 칩셋명 병합"""
    if not chipsets:
        return {}
    
    # SM으로 시작하는 칩셋과 JDM T618은 예외 처리
    excluded_patterns = [r'^sm', r'^jdm t618$']
    
    # 칩셋 그룹화
    chipset_groups = {}
    
    for chipset in chipsets:
        # SM으로 시작하는 칩셋과 JDM T618은 예외
        if any(re.match(pattern, chipset.lower()) for pattern in excluded_patterns):
            chipset_groups[chipset] = [chipset]
            continue
        
        # 정규화된 칩셋명
        normalized = normalize_chipset_name(chipset)
        
        # 그룹에 추가
        if normalized not in chipset_groups:
            chipset_groups[normalized] = []
        chipset_groups[normalized].append(chipset)
    
    # 각 그룹에서 가장 긴 칩셋명을 대표로 선택
    merged_chipsets = {}
    for normalized, group in chipset_groups.items():
        if len(group) == 1:
            merged_chipsets[group[0]] = group[0]
        else:
            # 가장 긴 칩셋명을 대표로 선택
            representative = max(group, key=len)
            for chipset in group:
                merged_chipsets[chipset] = representative
    
    return merged_chipsets

# ========== 유틸리티 함수 ==========

def convert_qdata_date(date_str):
    """
    Q-data 날짜 변환: 260209 → 2026-02-09
    """
    if pd.isna(date_str):
        return None
    try:
        date_str = str(int(date_str))  # 260209
        year = '20' + date_str[:2]     # 2026
        month = date_str[2:4]          # 02
        day = date_str[4:6]            # 09
        return f"{year}-{month}-{day}"
    except:
        return None

def read_qdata_excel(file_path):
    """
    Q-data 엑셀 파일 읽기 (DRM 처리 포함)
    - 9행부터 데이터 시작 (header=8)
    - F, M, P, Q, T, Z, AD, AR, BE, BF 열 읽기
    - DRM 파일 처리 (8가지 방법 fallback)
    """
    # 열 인덱스 (0부터 시작)
    # F=5, M=12, P=15, Q=16, T=19, Z=25, AD=29, AR=43, BE=50, BF=51
    usecols = [5, 12, 15, 16, 19, 25, 29, 43, 50, 51]
    df = None
    
    # DRM 처리 - 8가지 방법 시도
    methods = [
        # Method 1: openpyxl (기본)
        lambda: pd.read_excel(file_path, header=8, usecols=usecols, engine='openpyxl'),
        
        # Method 2: xlrd
        lambda: pd.read_excel(file_path, header=8, usecols=usecols, engine='xlrd'),
        
        # Method 3: pyxlsb (xlsb 형식)
        lambda: pd.read_excel(file_path, header=8, usecols=usecols, engine='pyxlsb'),
        
        # Method 4: 임시 파일로 복사 후 읽기
        lambda: read_via_temp_file(file_path, header=8, usecols=usecols),
        
        # Method 5: xlwings (Windows only)
        lambda: read_via_xlwings(file_path, header=8, usecols=usecols),
        
        # Method 6: win32com (Windows only)
        lambda: read_via_win32com(file_path, header=8, usecols=usecols),
        
        # Method 7: 메모리 스트림
        lambda: read_via_memory_stream(file_path, header=8, usecols=usecols),
        
        # Method 8: openpyxl 직접 사용
        lambda: read_via_openpyxl_direct(file_path, header=8, usecols=usecols)
    ]
    
    last_error = None
    for i, method in enumerate(methods, 1):
        try:
            df = method()
            if df is not None and not df.empty:
                print(f"✓ Q-data DRM 해제 성공 (방법 {i})")
                break
        except Exception as e:
            last_error = e
            continue
    
    if df is None:
        raise Exception(f"Q-data 엑셀 파일 읽기 실패 (DRM): {last_error}")
    
    # 컬럼 이름 매핑 (실제 헤더 이름과 관계없이 순서대로 매핑)
    df.columns = [
        'service_date',    # F열
        'process_type',    # M열
        'repair_name',     # P열
        'repair_detail',   # Q열
        'detail_content',  # T열
        'model_name',      # Z열
        'serial_number',   # AD열
        'log_id',          # AR열
        'sw_before',       # BE열
        'sw_after'         # BF열
    ]
    
    # 날짜 변환
    df['service_date'] = df['service_date'].apply(convert_qdata_date)
    
    # 빈 행 제거
    df = df.dropna(subset=['service_date', 'model_name'])
    
    return df

def read_via_temp_file(file_path, header, usecols):
    """Method 4: 임시 파일로 복사 후 읽기"""
    import tempfile
    import shutil
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp_path = tmp.name
        shutil.copy2(file_path, tmp_path)
        
        try:
            df = pd.read_excel(tmp_path, header=header, usecols=usecols, engine='openpyxl')
            return df
        finally:
            try:
                os.remove(tmp_path)
            except:
                pass

def read_via_xlwings(file_path, header, usecols):
    """Method 5: xlwings 사용 (Windows only)"""
    try:
        import xlwings as xw
        
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        ws = wb.sheets[0]
        
        # 데이터 읽기 (9행부터)
        data = ws.range('A9').expand().value
        
        wb.close()
        app.quit()
        
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # 필요한 열만 선택
        return df.iloc[:, usecols]
    except:
        raise

def read_via_win32com(file_path, header, usecols):
    """Method 6: win32com 사용 (Windows only)"""
    try:
        import win32com.client
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        
        wb = excel.Workbooks.Open(file_path)
        ws = wb.Worksheets(1)
        
        # 데이터 범위 읽기
        used_range = ws.UsedRange
        data = used_range.Value
        
        wb.Close(False)
        excel.Quit()
        
        df = pd.DataFrame(data[header+1:], columns=data[header])
        
        # 필요한 열만 선택
        return df.iloc[:, usecols]
    except:
        raise

def read_via_memory_stream(file_path, header, usecols):
    """Method 7: 메모리 스트림 사용"""
    import io
    
    with open(file_path, 'rb') as f:
        file_content = f.read()
    
    stream = io.BytesIO(file_content)
    return pd.read_excel(stream, header=header, usecols=usecols, engine='openpyxl')

def read_via_openpyxl_direct(file_path, header, usecols):
    """Method 8: openpyxl 직접 사용"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    # 데이터 읽기
    data = []
    for row in ws.iter_rows(min_row=header+2, values_only=True):
        selected_data = [row[i] for i in usecols if i < len(row)]
        if len(selected_data) == len(usecols):
            data.append(selected_data)
    
    wb.close()
    
    column_names = [
        'service_date', 'process_type', 'repair_name', 'repair_detail',
        'detail_content', 'model_name', 'serial_number', 'log_id',
        'sw_before', 'sw_after'
    ]
    
    return pd.DataFrame(data, columns=column_names)

def init_qdata_table():
    """Q-data 테이블 초기화"""
    conn = sqlite3.connect('voc_data.db')
    cursor = conn.cursor()
    
    # SQL 파일 실행
    with open('create_qdata_table.sql', 'r', encoding='utf-8') as f:
        sql_script = f.read()
        cursor.executescript(sql_script)
    
    conn.commit()
    conn.close()
# <<< qdata_backend.py 유틸리티 함수 끝 >>>>>

# ========== API 엔드포인트 ==========
@app.route('/')
def index():
    """메인 대시보드"""
    return render_template('index.html')

@app.route('/upload')
def upload_page():
    """업로드 페이지"""
    return render_template('upload.html')

@app.route('/statistics')
def statistics_page():
    """통계 페이지"""
    return render_template('statistics.html')

def process_voc_row(row, file_filename):
    """VOC 데이터 한 행 처리"""
    try:
        case_code = str(row.iloc[0]) if pd.notna(row.iloc[0]) else None  # A열
        title = str(row.iloc[7]) if pd.notna(row.iloc[7]) else None  # H열
        problem = str(row.iloc[12]) if pd.notna(row.iloc[12]) else None  # M열
        reproduction = str(row.iloc[13]) if pd.notna(row.iloc[13]) else None  # N열
        resolver = str(row.iloc[14]) if pd.notna(row.iloc[14]) else None  # O열
        resolve_option = str(row.iloc[17]) if pd.notna(row.iloc[17]) else None  # R열
        cause = str(row.iloc[20]) if pd.notna(row.iloc[20]) else None  # U열
        solution = str(row.iloc[21]) if pd.notna(row.iloc[21]) else None  # V열
        
        if not case_code:
            return None, None, None
        
        # 데이터 추출
        model_no = None  # 초기화
        watch_model = extract_watch_model(title)
        if watch_model:
            model_name = watch_model
        else:
            model_from_title = extract_model_from_title(title)
            model_no = extract_model_from_reproduction(reproduction)
            model_name = model_no or model_from_title
        
        # 모델명 매핑
        model_name = map_model_name(model_name)
        
        build_version = extract_build_version(reproduction)
        os_version = extract_os_version(reproduction)
        original_content = extract_original_content(reproduction)
        issue_type = detect_issue_type(problem, reproduction)
        
        # 칩셋 매핑
        chipset = get_chipset_for_model(model_name)
        
        # 3rd party 앱 감지
        search_text = f"{problem or ''} {original_content or ''}"
        third_party_app = detect_third_party_app(search_text)
        
        # 생성일자 추출 (사례코드에서 또는 파일명에서)
        created_date = None
        if case_code and case_code.startswith('P'):
            date_match = re.search(r'P(\d{6})', case_code)
            if date_match:
                date_str = date_match.group(1)
                try:
                    created_date = datetime.strptime(date_str, '%y%m%d').strftime('%Y-%m-%d')
                except:
                    pass
        else:
            # 사례코드 형식이 아닐 경우, 파일명에서 날짜 추출 시도
            date_match = re.search(r'(\d{8})', str(file_filename))
            if date_match:
                date_str = date_match.group(1)
                try:
                    created_date = datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
                except:
                    pass
        
        return case_code, {
            'title': title,
            'model_name': model_name,
            'model_no': model_no,
            'chipset': chipset,
            'build_version': build_version,
            'os_version': os_version,
            'issue_type': issue_type,
            'problem': problem,
            'original_content': original_content,
            'reproduction': reproduction,
            'resolver': resolver,
            'resolve_option': resolve_option,
            'cause': cause,
            'solution': solution,
            'third_party_app': third_party_app,
            'created_date': created_date
        }, chipset is None and model_name is not None
    
    except Exception as e:
        print(f"Row processing error: {str(e)}")
        return None, None, None

@app.route('/api/upload/internal_voc', methods=['POST'])
def upload_internal_voc():
    """사내 VOC 엑셀 업로드"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '파일이 없습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '파일이 선택되지 않았습니다.'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '엑셀 파일만 업로드 가능합니다.'}), 400
        
        print(f"파일 업로드 시작: {file.filename}")
        
        # DRM 처리 엑셀 파일 읽기
        try:
            df = read_excel_with_drm(file)
            print(f"엑셀 파일 읽기 성공: {len(df)}행")
        except Exception as e:
            print(f"엑셀 파일 읽기 실패: {str(e)}")
            return jsonify({'error': str(e)}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        success_count = 0
        error_count = 0
        unmapped_models = set()
        chunk_size = 1000  # 청크 크기
        total_rows = len(df)
        
        print(f"데이터 처리 시작: 총 {total_rows}행")
        
        # 청크 단위로 처리
        for chunk_start in range(0, total_rows, chunk_size):
            chunk_end = min(chunk_start + chunk_size, total_rows)
            chunk_df = df.iloc[chunk_start:chunk_end]
            
            print(f"청크 처리 중: {chunk_start + 1}-{chunk_end}/{total_rows}")
            
            for idx, row in chunk_df.iterrows():
                try:
                    case_code, voc_data, is_unmapped = process_voc_row(row, file.filename)
                    
                    if case_code is None:
                        continue
                    
                    if is_unmapped:
                        unmapped_models.add(voc_data['model_name'])
                    
                    # 중복 데이터 확인 및 처리
                    c.execute("SELECT id FROM internal_voc WHERE case_code = ?", (case_code,))
                    existing_record = c.fetchone()
                    
                    if existing_record:
                        # 기존 데이터가 있으면 모델명, U열(원인), V열(대책) 업데이트
                        c.execute("""UPDATE internal_voc
                                    SET model_name = ?, cause = ?, solution = ?, uploaded_date = ?
                                    WHERE case_code = ?""",
                                 (voc_data['model_name'], voc_data['cause'], voc_data['solution'], 
                                  datetime.now().strftime('%Y-%m-%d %H:%M:%S'), case_code))
                    else:
                        # 새 데이터이면 전체 삽입
                        c.execute("""INSERT INTO internal_voc 
                                    (case_code, title, model_name, model_no, chipset, build_version, 
                                     os_version, issue_type, problem, original_content, reproduction_path,
                                     resolver, resolve_option, cause, solution, third_party_app, 
                                     created_date, uploaded_date)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                 (case_code, voc_data['title'], voc_data['model_name'], voc_data['model_no'], 
                                  voc_data['chipset'], voc_data['build_version'], voc_data['os_version'], 
                                  voc_data['issue_type'], voc_data['problem'], voc_data['original_content'], 
                                  voc_data['reproduction'], voc_data['resolver'], voc_data['resolve_option'], 
                                  voc_data['cause'], voc_data['solution'], voc_data['third_party_app'],
                                  voc_data['created_date'], datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                    
                    success_count += 1
                    
                    # 100행마다 커밋하여 메모리 해제
                    if success_count % 100 == 0:
                        conn.commit()
                        print(f"진행률: {success_count}/{total_rows} ({success_count/total_rows*100:.1f}%)")
                
                except Exception as e:
                    error_count += 1
                    print(f"Row {idx} error: {str(e)}")
        
        conn.commit()
        conn.close()
        
        print(f"업로드 완료: 성공 {success_count}건, 실패 {error_count}건")
        
        # 결과 메시지 구성
        message = f'업로드 완료: {success_count}건 성공, {error_count}건 실패'
        if unmapped_models:
            message += f'\n칩셋 미매핑 모델: {len(unmapped_models)}개'
        
        return jsonify({
            'success': True,
            'message': message,
            'unmapped_models': list(unmapped_models) if unmapped_models else []
        })
    
    except Exception as e:
        print(f"업로드 실패: {str(e)}")
        return jsonify({'error': f'업로드 실패: {str(e)}'}), 500

@app.route('/api/upload/chipset_mapping', methods=['POST'])
def upload_chipset_mapping():
    """칩셋 매핑 파일 업로드"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '파일이 없습니다.'}), 400
        
        file = request.files['file']
        
        # DRM 처리 엑셀 파일 읽기
        try:
            df = read_excel_with_drm(file)
        except Exception as e:
            return jsonify({'error': str(e)}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        success_count = 0
        update_count = 0
        duplicate_count = 0
        
        # 모델명과 칩셋명 저장
        model_chipset_pairs = []
        for _, row in df.iterrows():
            model_name = str(row.iloc[0]).strip()
            chipset = str(row.iloc[1]).strip()
            
            if model_name and chipset:
                model_chipset_pairs.append((model_name, chipset))
        
        # 칩셋명 병합
        all_chipsets = [chipset for _, chipset in model_chipset_pairs]
        merged_chipsets = merge_similar_chipsets(all_chipsets)
        
        # 병합된 칩셋명으로 매핑 저장 및 업데이트
        for model_name, chipset in model_chipset_pairs:
            # 병합된 칩셋명 사용
            final_chipset = merged_chipsets.get(chipset, chipset)
            
            if chipset != final_chipset:
                print(f"칩셋명 병합: {chipset} -> {final_chipset}")
            
            try:
                c.execute("INSERT INTO chipset_mapping (model_name, chipset) VALUES (?, ?)",
                         (model_name, final_chipset))
                success_count += 1
                
                # 관련 VOC 데이터의 칩셋도 업데이트
                c.execute("UPDATE internal_voc SET chipset = ? WHERE model_name = ?",
                         (final_chipset, model_name))
                
                print(f"칩셋 매핑: {model_name} -> {final_chipset}")
                
            except sqlite3.IntegrityError as e:
                if "UNIQUE constraint failed" in str(e):
                    # 중복 모델명이면 칩셋만 업데이트
                    c.execute("UPDATE chipset_mapping SET chipset = ? WHERE model_name = ?",
                             (final_chipset, model_name))
                    
                    # 관련 VOC 데이터의 칩셋도 업데이트
                    c.execute("UPDATE internal_voc SET chipset = ? WHERE model_name = ?",
                             (final_chipset, model_name))
                    
                    update_count += 1
                    duplicate_count += 1
                    print(f"칩셋 업데이트: {model_name} -> {final_chipset}")
                else:
                    raise e
        
        conn.commit()
        conn.close()
        
        message = f'{success_count}개의 칩셋 매핑이 등록되었습니다.'
        if update_count > 0:
            message += f' {update_count}개가 업데이트되었습니다.'
        
        return jsonify({
            'success': True,
            'message': message
        })
    
    except Exception as e:
        return jsonify({'error': f'업로드 실패: {str(e)}'}), 500

@app.route('/api/upload/app_keywords', methods=['POST'])
def upload_app_keywords():
    """3rd party 앱 키워드 파일 업로드"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '파일이 없습니다.'}), 400
        
        file = request.files['file']
        
        # DRM 처리 엑셀 파일 읽기
        try:
            df = read_excel_with_drm(file)
        except Exception as e:
            return jsonify({'error': str(e)}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 기존 데이터 삭제
        c.execute("DELETE FROM app_keywords")
        
        success_count = 0
        for _, row in df.iterrows():
            app_name = str(row.iloc[0]).strip()
            keywords = str(row.iloc[1]).strip()
            
            if app_name and keywords:
                c.execute("INSERT INTO app_keywords (app_name, keywords) VALUES (?, ?)",
                         (app_name, keywords))
                success_count += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{success_count}개의 앱 키워드가 등록되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'업로드 실패: {str(e)}'}), 500

@app.route('/api/dashboard/daily')
def get_daily_dashboard():
    """일일 대시보드 데이터"""
    try:
        conn = sqlite3.connect('voc_data.db')
        
        # 어제 날짜
        yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        
        # 전일 VOC 건수
        query_daily = """
            SELECT COUNT(*) as count
            FROM internal_voc
            WHERE DATE(created_date) = ?
        """
        df_daily = pd.read_sql_query(query_daily, conn, params=(yesterday,))
        daily_count = df_daily.iloc[0]['count']
        
        # 모델별 Top 10
        query_top10 = """
            SELECT model_name, COUNT(*) as count
            FROM internal_voc
            WHERE DATE(created_date) = ? AND model_name IS NOT NULL
            GROUP BY model_name
            ORDER BY count DESC
            LIMIT 10
        """
        df_top10 = pd.read_sql_query(query_top10, conn, params=(yesterday,))
        
        conn.close()
        
        return jsonify({
            'yesterday_date': yesterday,
            'daily_count': int(daily_count),
            'top10_models': df_top10.to_dict('records')
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics/model')
def get_model_statistics():
    """휴대폰 모델별 통계"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = sqlite3.connect('voc_data.db')
        
        query = """
            SELECT model_name, COUNT(*) as count
            FROM internal_voc
            WHERE model_name IS NOT NULL
        """
        params = []
        
        if start_date and end_date:
            query += " AND DATE(created_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        
        query += " GROUP BY model_name ORDER BY count DESC"
        
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        
        return jsonify(df.to_dict('records'))
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics/weekly')
def get_weekly_statistics():
    """주별 통계"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = sqlite3.connect('voc_data.db')
        
        query = """
            SELECT 
                strftime('%Y-%W', created_date) as week,
                COUNT(*) as count
            FROM internal_voc
            WHERE created_date IS NOT NULL
        """
        params = []
        
        if start_date and end_date:
            query += " AND DATE(created_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        
        query += " GROUP BY week ORDER BY week"
        
        df = pd.read_sql_query(query, conn, params=params)
        
        # 메모 정보 조회
        c = conn.cursor()
        c.execute("SELECT week, memo FROM weekly_memos")
        memos = {row[0]: row[1] for row in c.fetchall()}
        
        conn.close()
        
        # 메모 정보를 주별 데이터에 추가
        result = df.to_dict('records')
        for item in result:
            item['memo'] = memos.get(item['week'], '')
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics/monthly')
def get_monthly_statistics():
    """월별 통계"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = sqlite3.connect('voc_data.db')
        
        query = """
            SELECT 
                strftime('%Y-%m', created_date) as month,
                COUNT(*) as count
            FROM internal_voc
            WHERE created_date IS NOT NULL
        """
        params = []
        
        if start_date and end_date:
            query += " AND DATE(created_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        
        query += " GROUP BY month ORDER BY month"
        
        df = pd.read_sql_query(query, conn, params=params)
        
        # 메모 정보 조회
        c = conn.cursor()
        c.execute("SELECT month, memo FROM monthly_memos")
        memos = {row[0]: row[1] for row in c.fetchall()}
        
        conn.close()
        
        # 메모 정보를 월별 데이터에 추가
        result = df.to_dict('records')
        for item in result:
            item['memo'] = memos.get(item['month'], '')
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics/chipset')
def get_chipset_statistics():
    """칩셋별 통계"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = sqlite3.connect('voc_data.db')
        
        query = """
            SELECT chipset, COUNT(*) as count
            FROM internal_voc
            WHERE chipset IS NOT NULL
        """
        params = []
        
        if start_date and end_date:
            query += " AND DATE(created_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        
        query += " GROUP BY chipset ORDER BY count DESC"
        
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        
        return jsonify(df.to_dict('records'))
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics/app')
def get_app_statistics():
    """3rd party 앱별 통계"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = sqlite3.connect('voc_data.db')
        
        query = """
            SELECT third_party_app, COUNT(*) as count
            FROM internal_voc
            WHERE third_party_app IS NOT NULL
        """
        params = []
        
        if start_date and end_date:
            query += " AND DATE(created_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        
        query += " GROUP BY third_party_app ORDER BY count DESC"
        
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        
        return jsonify(df.to_dict('records'))
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/voc/<int:voc_id>')
def get_voc_detail(voc_id):
    """VOC 상세 정보"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        c.execute("SELECT * FROM internal_voc WHERE id = ?", (voc_id,))
        voc = c.fetchone()
        
        if not voc:
            return jsonify({'error': 'VOC를 찾을 수 없습니다.'}), 404
        
        columns = [desc[0] for desc in c.description]
        voc_dict = dict(zip(columns, voc))
        
        # 댓글 조회
        c.execute("""
            SELECT id, comment, created_date 
            FROM comments 
            WHERE voc_id = ? AND voc_type = 'internal'
            ORDER BY created_date DESC
        """, (voc_id,))
        comments = c.fetchall()
        
        conn.close()
        
        voc_dict['comments'] = [
            {'id': c[0], 'comment': c[1], 'created_date': c[2]}
            for c in comments
        ]
        
        return jsonify(voc_dict)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/voc/<int:voc_id>/comment', methods=['POST'])
def add_comment(voc_id):
    """댓글 추가"""
    try:
        data = request.get_json()
        comment = data.get('comment', '').strip()
        
        if not comment:
            return jsonify({'error': '댓글 내용을 입력해주세요.'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        c.execute("""
            INSERT INTO comments (voc_id, voc_type, comment, created_date)
            VALUES (?, 'internal', ?, ?)
        """, (voc_id, comment, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        
        conn.commit()
        comment_id = c.lastrowid
        conn.close()
        
        return jsonify({
            'success': True,
            'comment_id': comment_id,
            'message': '댓글이 추가되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel')
def export_to_excel():
    """엑셀로 내보내기"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = sqlite3.connect('voc_data.db')
        
        query = "SELECT * FROM internal_voc"
        params = []
        
        if start_date and end_date:
            query += " WHERE DATE(created_date) BETWEEN ? AND ?"
            params = [start_date, end_date]
        
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        
        # 엑셀 파일 생성
        filename = f"voc_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        df.to_excel(filepath, index=False)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reset/data', methods=['POST'])
def reset_voc_data():
    """기존 업로드 데이터 초기화"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # VOC 데이터 삭제 건수 확인
        c.execute("SELECT COUNT(*) FROM internal_voc")
        voc_count = c.fetchone()[0]
        
        c.execute("SELECT COUNT(*) FROM comments")
        comment_count = c.fetchone()[0]
        
        # VOC 데이터 삭제
        c.execute("DELETE FROM internal_voc")
        
        # 댓글 데이터 삭제
        c.execute("DELETE FROM comments")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'데이터 초기화 완료: VOC {voc_count}건, 댓글 {comment_count}건 삭제'
        })
    
    except Exception as e:
        return jsonify({'error': f'데이터 초기화 실패: {str(e)}'}), 500

@app.route('/api/unmapped-models')
def get_unmapped_models():
    """칩셋 미매핑 모델명 조회"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 모델명이 있지만 칩셋이 없는 데이터 조회
        c.execute("""
            SELECT DISTINCT model_name, COUNT(*) as count
            FROM internal_voc 
            WHERE model_name IS NOT NULL 
            AND model_name != ''
            AND (chipset IS NULL OR chipset = '')
            GROUP BY model_name
            ORDER BY count DESC
        """)
        unmapped_models = c.fetchall()
        
        # 전체 모델명 수와 미매핑 모델명 수
        c.execute("SELECT COUNT(DISTINCT model_name) FROM internal_voc WHERE model_name IS NOT NULL AND model_name != ''")
        total_models = c.fetchone()[0]
        
        conn.close()
        
        return jsonify({
            'success': True,
            'unmapped_models': [{'model_name': row[0], 'count': row[1]} for row in unmapped_models],
            'total_unmapped': len(unmapped_models),
            'total_models': total_models
        })
    
    except Exception as e:
        return jsonify({'error': f'조회 실패: {str(e)}'}), 500

@app.route('/api/chipset-mapping/add', methods=['POST'])
def add_chipset_mapping():
    """개별 칩셋 매핑 추가"""
    try:
        data = request.get_json()
        model_name = data.get('model_name', '').strip()
        chipset = data.get('chipset', '').strip()
        
        if not model_name or not chipset:
            return jsonify({'error': '모델명과 칩셋명을 모두 입력해주세요.'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 기존 매핑 확인
        c.execute("SELECT id FROM chipset_mapping WHERE model_name = ?", (model_name,))
        existing = c.fetchone()
        
        if existing:
            # 기존 매핑이 있으면 업데이트
            c.execute("UPDATE chipset_mapping SET chipset = ? WHERE model_name = ?", 
                     (chipset, model_name))
            action = '업데이트'
        else:
            # 새 매핑 추가
            c.execute("INSERT INTO chipset_mapping (model_name, chipset) VALUES (?, ?)", 
                     (model_name, chipset))
            action = '추가'
        
        # 관련 VOC 데이터의 칩셋도 업데이트
        c.execute("UPDATE internal_voc SET chipset = ? WHERE model_name = ?", 
                 (chipset, model_name))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{model_name} 모델의 칩셋을 {chipset}(으)로 {action}했습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'추가 실패: {str(e)}'}), 500

@app.route('/api/chipset-mapping/batch', methods=['POST'])
def add_chipset_mapping_batch():
    """여러 칩셋 매핑 일괄 추가"""
    try:
        data = request.get_json()
        mappings = data.get('mappings', [])
        
        if not mappings:
            return jsonify({'error': '매핑 데이터가 없습니다.'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        success_count = 0
        error_count = 0
        errors = []
        
        for mapping in mappings:
            try:
                model_name = mapping.get('model_name', '').strip()
                chipset = mapping.get('chipset', '').strip()
                
                if not model_name or not chipset:
                    errors.append(f'모델명 또는 칩셋명이 누락됨: {mapping}')
                    error_count += 1
                    continue
                
                # 기존 매핑 확인
                c.execute("SELECT id FROM chipset_mapping WHERE model_name = ?", (model_name,))
                existing = c.fetchone()
                
                if existing:
                    c.execute("UPDATE chipset_mapping SET chipset = ? WHERE model_name = ?", 
                             (chipset, model_name))
                else:
                    c.execute("INSERT INTO chipset_mapping (model_name, chipset) VALUES (?, ?)", 
                             (model_name, chipset))
                
                # 관련 VOC 데이터의 칩셋도 업데이트
                c.execute("UPDATE internal_voc SET chipset = ? WHERE model_name = ?", 
                         (chipset, model_name))
                
                success_count += 1
                
            except Exception as e:
                errors.append(f'{mapping.get("model_name", "Unknown")}: {str(e)}')
                error_count += 1
        
        conn.commit()
        conn.close()
        
        message = f'일괄 처리 완료: {success_count}건 성공'
        if error_count > 0:
            message += f', {error_count}건 실패'
        
        return jsonify({
            'success': True,
            'message': message,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:5]  # 최대 5개 오류만 반환
        })
    
    except Exception as e:
        return jsonify({'error': f'일괄 추가 실패: {str(e)}'}), 500

@app.route('/api/update/created_dates', methods=['POST'])
def update_created_dates():
    """생성일자 업데이트 (파일명에서 날짜 추출)"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # created_date가 NULL인 데이터 조회
        c.execute("""
            SELECT id, case_code, uploaded_date 
            FROM internal_voc 
            WHERE created_date IS NULL OR created_date = ''
        """)
        records = c.fetchall()
        
        updated_count = 0
        for record in records:
            voc_id, case_code, uploaded_date = record
            
            # uploaded_date에서 파일명 추출 (업로드 시간 기준으로 파일명에 날짜가 포함된 경우)
            date_match = re.search(r'(\d{8})', uploaded_date)
            if date_match:
                date_str = date_match.group(1)
                try:
                    created_date = datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
                    c.execute("UPDATE internal_voc SET created_date = ? WHERE id = ?", 
                             (created_date, voc_id))
                    updated_count += 1
                except:
                    pass
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{updated_count}개의 데이터 생성일자를 업데이트했습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'업데이트 실패: {str(e)}'}), 500

@app.route('/api/monthly-memos', methods=['GET'])
def get_monthly_memos():
    """전체 월별 메모 조회"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        c.execute("""
            SELECT month, memo, created_date, updated_date
            FROM monthly_memos
            ORDER BY month DESC
        """)
        memos = c.fetchall()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'memos': [
                {
                    'month': row[0],
                    'memo': row[1],
                    'created_date': row[2],
                    'updated_date': row[3]
                }
                for row in memos
            ]
        })
    
    except Exception as e:
        return jsonify({'error': f'조회 실패: {str(e)}'}), 500

@app.route('/api/monthly-memos', methods=['POST'])
def add_monthly_memo():
    """월별 메모 추가"""
    try:
        data = request.get_json()
        month = data.get('month', '').strip()
        memo = data.get('memo', '').strip()
        
        if not month or not memo:
            return jsonify({'error': '월과 메모를 모두 입력해주세요.'}), 400
        
        # 월 형식 검증 (YYYY-MM)
        if not re.match(r'^\d{4}-\d{2}$', month):
            return jsonify({'error': '월 형식이 올바르지 않습니다. (YYYY-MM)'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 기존 메모 확인
        c.execute("SELECT id FROM monthly_memos WHERE month = ?", (month,))
        existing = c.fetchone()
        
        if existing:
            conn.close()
            return jsonify({'error': f'{month}월에 이미 메모가 있습니다. 수정을 사용해주세요.'}), 400
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        c.execute("""
            INSERT INTO monthly_memos (month, memo, created_date, updated_date)
            VALUES (?, ?, ?, ?)
        """, (month, memo, now, now))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{month}월에 메모가 추가되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'추가 실패: {str(e)}'}), 500

@app.route('/api/monthly-memos/<month>', methods=['PUT'])
def update_monthly_memo(month):
    """월별 메모 수정"""
    try:
        data = request.get_json()
        memo = data.get('memo', '').strip()
        
        if not memo:
            return jsonify({'error': '메모를 입력해주세요.'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 메모 존재 확인
        c.execute("SELECT id FROM monthly_memos WHERE month = ?", (month,))
        existing = c.fetchone()
        
        if not existing:
            conn.close()
            return jsonify({'error': f'{month}월에 메모가 없습니다.'}), 404
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        c.execute("""
            UPDATE monthly_memos
            SET memo = ?, updated_date = ?
            WHERE month = ?
        """, (memo, now, month))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{month}월 메모가 수정되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'수정 실패: {str(e)}'}), 500

@app.route('/api/monthly-memos/<month>', methods=['DELETE'])
def delete_monthly_memo(month):
    """월별 메모 삭제"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 메모 존재 확인
        c.execute("SELECT id FROM monthly_memos WHERE month = ?", (month,))
        existing = c.fetchone()
        
        if not existing:
            conn.close()
            return jsonify({'error': f'{month}월에 메모가 없습니다.'}), 404
        
        c.execute("DELETE FROM monthly_memos WHERE month = ?", (month,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{month}월 메모가 삭제되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'삭제 실패: {str(e)}'}), 500

@app.route('/api/weekly-memos', methods=['GET'])
def get_weekly_memos():
    """전체 주별 메모 조회"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        c.execute("""
            SELECT week, memo, created_date, updated_date
            FROM weekly_memos
            ORDER BY week DESC
        """)
        memos = c.fetchall()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'memos': [
                {
                    'week': row[0],
                    'memo': row[1],
                    'created_date': row[2],
                    'updated_date': row[3]
                }
                for row in memos
            ]
        })
    
    except Exception as e:
        return jsonify({'error': f'조회 실패: {str(e)}'}), 500

@app.route('/api/weekly-memos', methods=['POST'])
def add_weekly_memo():
    """주별 메모 추가"""
    try:
        data = request.get_json()
        week = data.get('week', '').strip()
        memo = data.get('memo', '').strip()
        
        if not week or not memo:
            return jsonify({'error': '주와 메모를 모두 입력해주세요.'}), 400
        
        # 주 형식 검증 (YYYY-WW)
        if not re.match(r'^\d{4}-\d{2}$', week):
            return jsonify({'error': '주 형식이 올바르지 않습니다. (YYYY-WW)'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 기존 메모 확인
        c.execute("SELECT id FROM weekly_memos WHERE week = ?", (week,))
        existing = c.fetchone()
        
        if existing:
            conn.close()
            return jsonify({'error': f'{week}주에 이미 메모가 있습니다. 수정을 사용해주세요.'}), 400
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        c.execute("""
            INSERT INTO weekly_memos (week, memo, created_date, updated_date)
            VALUES (?, ?, ?, ?)
        """, (week, memo, now, now))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{week}주에 메모가 추가되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'추가 실패: {str(e)}'}), 500

@app.route('/api/weekly-memos/<week>', methods=['PUT'])
def update_weekly_memo(week):
    """주별 메모 수정"""
    try:
        data = request.get_json()
        memo = data.get('memo', '').strip()
        
        if not memo:
            return jsonify({'error': '메모를 입력해주세요.'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 메모 존재 확인
        c.execute("SELECT id FROM weekly_memos WHERE week = ?", (week,))
        existing = c.fetchone()
        
        if not existing:
            conn.close()
            return jsonify({'error': f'{week}주에 메모가 없습니다.'}), 404
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        c.execute("""
            UPDATE weekly_memos
            SET memo = ?, updated_date = ?
            WHERE week = ?
        """, (memo, now, week))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{week}주 메모가 수정되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'수정 실패: {str(e)}'}), 500

@app.route('/api/weekly-memos/<week>', methods=['DELETE'])
def delete_weekly_memo(week):
    """주별 메모 삭제"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 메모 존재 확인
        c.execute("SELECT id FROM weekly_memos WHERE week = ?", (week,))
        existing = c.fetchone()
        
        if not existing:
            conn.close()
            return jsonify({'error': f'{week}주에 메모가 없습니다.'}), 404
        
        c.execute("DELETE FROM weekly_memos WHERE week = ?", (week,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{week}주 메모가 삭제되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'삭제 실패: {str(e)}'}), 500

@app.route('/api/model-monthly-memos/<model_name>', methods=['GET'])
def get_model_monthly_memos(model_name):
    """특정 모델의 월별 메모 조회"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        c.execute("""
            SELECT month, memo, created_date, updated_date
            FROM model_monthly_memos
            WHERE model_name = ?
            ORDER BY month DESC
        """, (model_name,))
        memos = c.fetchall()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'memos': [
                {
                    'month': row[0],
                    'memo': row[1],
                    'created_date': row[2],
                    'updated_date': row[3]
                }
                for row in memos
            ]
        })
    
    except Exception as e:
        return jsonify({'error': f'조회 실패: {str(e)}'}), 500

@app.route('/api/model-monthly-memos', methods=['POST'])
def add_model_monthly_memo():
    """모델별 월별 메모 추가"""
    try:
        data = request.get_json()
        model_name = data.get('model_name', '').strip()
        month = data.get('month', '').strip()
        memo = data.get('memo', '').strip()
        
        if not model_name or not month or not memo:
            return jsonify({'error': '모델명, 월, 메모를 모두 입력해주세요.'}), 400
        
        # 월 형식 검증 (YYYY-MM)
        if not re.match(r'^\d{4}-\d{2}$', month):
            return jsonify({'error': '월 형식이 올바르지 않습니다. (YYYY-MM)'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 기존 메모 확인
        c.execute("SELECT id FROM model_monthly_memos WHERE model_name = ? AND month = ?", 
                 (model_name, month))
        existing = c.fetchone()
        
        if existing:
            conn.close()
            return jsonify({'error': f'{model_name} 모델의 {month}월에 이미 메모가 있습니다. 수정을 사용해주세요.'}), 400
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        c.execute("""
            INSERT INTO model_monthly_memos (model_name, month, memo, created_date, updated_date)
            VALUES (?, ?, ?, ?, ?)
        """, (model_name, month, memo, now, now))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{model_name} 모델의 {month}월에 메모가 추가되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'추가 실패: {str(e)}'}), 500

@app.route('/api/model-monthly-memos/<model_name>/<month>', methods=['PUT'])
def update_model_monthly_memo(model_name, month):
    """모델별 월별 메모 수정"""
    try:
        data = request.get_json()
        memo = data.get('memo', '').strip()
        
        if not memo:
            return jsonify({'error': '메모를 입력해주세요.'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 메모 존재 확인
        c.execute("SELECT id FROM model_monthly_memos WHERE model_name = ? AND month = ?", 
                 (model_name, month))
        existing = c.fetchone()
        
        if not existing:
            conn.close()
            return jsonify({'error': f'{model_name} 모델의 {month}월에 메모가 없습니다.'}), 404
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        c.execute("""
            UPDATE model_monthly_memos
            SET memo = ?, updated_date = ?
            WHERE model_name = ? AND month = ?
        """, (memo, now, model_name, month))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{model_name} 모델의 {month}월 메모가 수정되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'수정 실패: {str(e)}'}), 500

@app.route('/api/model-monthly-memos/<model_name>/<month>', methods=['DELETE'])
def delete_model_monthly_memo(model_name, month):
    """모델별 월별 메모 삭제"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 메모 존재 확인
        c.execute("SELECT id FROM model_monthly_memos WHERE model_name = ? AND month = ?", 
                 (model_name, month))
        existing = c.fetchone()
        
        if not existing:
            conn.close()
            return jsonify({'error': f'{model_name} 모델의 {month}월에 메모가 없습니다.'}), 404
        
        c.execute("DELETE FROM model_monthly_memos WHERE model_name = ? AND month = ?", 
                 (model_name, month))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{model_name} 모델의 {month}월 메모가 삭제되었습니다.'
        })
    
    except Exception as e:
        return jsonify({'error': f'삭제 실패: {str(e)}'}), 500

@app.route('/api/statistics/model/<model_name>/monthly')
def get_model_monthly_statistics(model_name):
    """특정 모델의 월별 통계"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = sqlite3.connect('voc_data.db')
        
        query = """
            SELECT 
                strftime('%Y-%m', created_date) as month,
                COUNT(*) as count
            FROM internal_voc
            WHERE model_name = ? AND created_date IS NOT NULL
        """
        params = [model_name]
        
        if start_date and end_date:
            query += " AND DATE(created_date) BETWEEN ? AND ?"
            params.extend([start_date, end_date])
        
        query += " GROUP BY month ORDER BY month"
        
        df = pd.read_sql_query(query, conn, params=params)
        
        # 메모 정보 조회
        c = conn.cursor()
        c.execute("SELECT month, memo FROM model_monthly_memos WHERE model_name = ?", (model_name,))
        memos = {row[0]: row[1] for row in c.fetchall()}
        
        conn.close()
        
        # 메모 정보를 월별 데이터에 추가
        result = df.to_dict('records')
        for item in result:
            item['memo'] = memos.get(item['month'], '')
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics/models/monthly', methods=['POST'])
def get_models_monthly_statistics():
    """여러 모델의 월별 통계"""
    try:
        data = request.get_json()
        model_names = data.get('model_names', [])
        
        if not model_names:
            return jsonify({'error': '모델명을 선택해주세요.'}), 400
        
        if len(model_names) > 10:
            return jsonify({'error': '최대 10개 모델까지만 선택 가능합니다.'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        
        # 각 모델의 월별 통계 조회
        result = {}
        for model_name in model_names:
            query = """
                SELECT 
                    strftime('%Y-%m', created_date) as month,
                    COUNT(*) as count
                FROM internal_voc
                WHERE model_name = ? AND created_date IS NOT NULL
                GROUP BY month ORDER BY month
            """
            df = pd.read_sql_query(query, conn, params=(model_name,))
            
            # 메모 정보 조회
            c = conn.cursor()
            c.execute("SELECT month, memo FROM model_monthly_memos WHERE model_name = ?", (model_name,))
            memos = {row[0]: row[1] for row in c.fetchall()}
            
            # 메모 정보를 월별 데이터에 추가
            model_data = df.to_dict('records')
            for item in model_data:
                item['memo'] = memos.get(item['month'], '')
            
            result[model_name] = model_data
        
        conn.close()
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/chipset/merge', methods=['POST'])
def merge_chipsets():
    """기존 칩셋명 병합"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 모든 칩셋명 조회
        c.execute("SELECT DISTINCT chipset FROM internal_voc WHERE chipset IS NOT NULL AND chipset != ''")
        all_chipsets = [row[0] for row in c.fetchall()]
        
        # 칩셋명 병합
        merged_chipsets = merge_similar_chipsets(all_chipsets)
        
        # 병합된 칩셋명으로 업데이트
        updated_count = 0
        for original_chipset, merged_chipset in merged_chipsets.items():
            if original_chipset != merged_chipset:
                c.execute("UPDATE internal_voc SET chipset = ? WHERE chipset = ?",
                         (merged_chipset, original_chipset))
                updated_count += c.rowcount
        
        # 칩셋 매핑 테이블도 업데이트
        c.execute("SELECT model_name, chipset FROM chipset_mapping")
        mappings = c.fetchall()
        for model_name, chipset in mappings:
            merged_chipset = merged_chipsets.get(chipset, chipset)
            if chipset != merged_chipset:
                c.execute("UPDATE chipset_mapping SET chipset = ? WHERE model_name = ?",
                         (merged_chipset, model_name))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{updated_count}개의 칩셋명이 병합되었습니다.'
        })
    except Exception as e:
        return jsonify({'error': f'병합 실패: {str(e)}'}), 500

@app.route('/api/chipset/rename', methods=['POST'])
def rename_chipset():
    """특정 칩셋명 변경"""
    try:
        data = request.get_json()
        old_chipset = data.get('old_chipset', '').strip()
        new_chipset = data.get('new_chipset', '').strip()
        
        if not old_chipset or not new_chipset:
            return jsonify({'error': '기존 칩셋명과 새 칩셋명을 모두 입력해주세요.'}), 400
        
        if old_chipset == new_chipset:
            return jsonify({'error': '기존 칩셋명과 새 칩셋명이 같습니다.'}), 400
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # internal_voc 테이블 업데이트
        c.execute("UPDATE internal_voc SET chipset = ? WHERE chipset = ?", (new_chipset, old_chipset))
        voc_updated_count = c.rowcount
        
        # chipset_mapping 테이블 업데이트
        c.execute("UPDATE chipset_mapping SET chipset = ? WHERE chipset = ?", (new_chipset, old_chipset))
        mapping_updated_count = c.rowcount
        
        conn.commit()
        conn.close()
        
        total_updated = voc_updated_count + mapping_updated_count
        return jsonify({
            'success': True,
            'message': f'{old_chipset}을 {new_chipset}(으)로 변경했습니다. (VOC: {voc_updated_count}건, 매핑: {mapping_updated_count}건)'
        })
    except Exception as e:
        return jsonify({'error': f'변경 실패: {str(e)}'}), 500

@app.route('/api/model/update-watch', methods=['POST'])
def update_watch_models():
    """기존 데이터의 모델명을 '워치' 단어로 업데이트"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 모든 데이터 조회
        c.execute("SELECT id, title, model_name FROM internal_voc WHERE title IS NOT NULL")
        records = c.fetchall()
        
        updated_count = 0
        for record in records:
            voc_id, title, current_model_name = record
            
            # '워치' 단어 추출
            watch_model = extract_watch_model(title)
            
            if watch_model and watch_model != current_model_name:
                # 모델명 업데이트
                c.execute("UPDATE internal_voc SET model_name = ? WHERE id = ?", (watch_model, voc_id))
                updated_count += 1
                print(f"모델명 업데이트: {current_model_name} -> {watch_model}")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{updated_count}개의 모델명이 업데이트되었습니다.'
        })
    except Exception as e:
        return jsonify({'error': f'업데이트 실패: {str(e)}'}), 500

@app.route('/api/model/update-mapping', methods=['POST'])
def update_model_mapping():
    """기존 데이터의 모델명을 매핑 규칙에 따라 업데이트"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 모든 데이터 조회
        c.execute("SELECT id, model_name FROM internal_voc WHERE model_name IS NOT NULL")
        records = c.fetchall()
        
        updated_count = 0
        for record in records:
            voc_id, current_model_name = record
            
            # 모델명 매핑
            new_model_name = map_model_name(current_model_name)
            
            if new_model_name != current_model_name:
                # 모델명 업데이트
                c.execute("UPDATE internal_voc SET model_name = ? WHERE id = ?", (new_model_name, voc_id))
                updated_count += 1
                print(f"모델명 업데이트: {current_model_name} -> {new_model_name}")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{updated_count}개의 모델명이 업데이트되었습니다.'
        })
    except Exception as e:
        return jsonify({'error': f'업데이트 실패: {str(e)}'}), 500

@app.route('/api/memos/backup', methods=['POST'])
def backup_memos():
    """메모 백업"""
    try:
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        # 월별 메모 백업
        c.execute("SELECT month, memo, created_date, updated_date FROM monthly_memos")
        monthly_memos = c.fetchall()
        
        # 주별 메모 백업
        c.execute("SELECT week, memo, created_date, updated_date FROM weekly_memos")
        weekly_memos = c.fetchall()
        
        # 모델별 월별 메모 백업
        c.execute("SELECT model_name, month, memo, created_date, updated_date FROM model_monthly_memos")
        model_monthly_memos = c.fetchall()
        
        conn.close()
        
        # 백업 파일 생성
        backup_data = {
            'monthly_memos': [
                {
                    'month': row[0],
                    'memo': row[1],
                    'created_date': row[2],
                    'updated_date': row[3]
                }
                for row in monthly_memos
            ],
            'weekly_memos': [
                {
                    'week': row[0],
                    'memo': row[1],
                    'created_date': row[2],
                    'updated_date': row[3]
                }
                for row in weekly_memos
            ],
            'model_monthly_memos': [
                {
                    'model_name': row[0],
                    'month': row[1],
                    'memo': row[2],
                    'created_date': row[3],
                    'updated_date': row[4]
                }
                for row in model_monthly_memos
            ],
            'backup_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # 백업 파일 저장
        backup_filename = f"memos_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        backup_filepath = os.path.join(app.config['UPLOAD_FOLDER'], backup_filename)
        
        with open(backup_filepath, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'메모 백업 완료: 월별 {len(monthly_memos)}개, 주별 {len(weekly_memos)}개, 모델별 {len(model_monthly_memos)}개',
            'backup_file': backup_filename
        })
    except Exception as e:
        return jsonify({'error': f'백업 실패: {str(e)}'}), 500

@app.route('/api/memos/restore', methods=['POST'])
def restore_memos():
    """메모 복구"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '백업 파일이 없습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '백업 파일이 선택되지 않았습니다.'}), 400
        
        if not file.filename.endswith('.json'):
            return jsonify({'error': 'JSON 파일만 업로드 가능합니다.'}), 400
        
        # 백업 파일 읽기
        backup_data = json.load(file)
        
        conn = sqlite3.connect('voc_data.db')
        c = conn.cursor()
        
        restored_count = 0
        
        # 월별 메모 복구
        if 'monthly_memos' in backup_data:
            for memo in backup_data['monthly_memos']:
                try:
                    c.execute("""
                        INSERT OR REPLACE INTO monthly_memos (month, memo, created_date, updated_date)
                        VALUES (?, ?, ?, ?)
                    """, (memo['month'], memo['memo'], memo['created_date'], memo['updated_date']))
                    restored_count += 1
                except Exception as e:
                    print(f"월별 메모 복구 실패: {memo['month']}, {str(e)}")
        
        # 주별 메모 복구
        if 'weekly_memos' in backup_data:
            for memo in backup_data['weekly_memos']:
                try:
                    c.execute("""
                        INSERT OR REPLACE INTO weekly_memos (week, memo, created_date, updated_date)
                        VALUES (?, ?, ?, ?)
                    """, (memo['week'], memo['memo'], memo['created_date'], memo['updated_date']))
                    restored_count += 1
                except Exception as e:
                    print(f"주별 메모 복구 실패: {memo['week']}, {str(e)}")
        
        # 모델별 월별 메모 복구
        if 'model_monthly_memos' in backup_data:
            for memo in backup_data['model_monthly_memos']:
                try:
                    c.execute("""
                        INSERT OR REPLACE INTO model_monthly_memos (model_name, month, memo, created_date, updated_date)
                        VALUES (?, ?, ?, ?, ?)
                    """, (memo['model_name'], memo['month'], memo['memo'], memo['created_date'], memo['updated_date']))
                    restored_count += 1
                except Exception as e:
                    print(f"모델별 월별 메모 복구 실패: {memo['model_name']}, {memo['month']}, {str(e)}")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'메모 복구 완료: {restored_count}개'
        })
    except Exception as e:
        return jsonify({'error': f'복구 실패: {str(e)}'}), 500

# ==========(qdata_backend.py에서 복사) API 엔드포인트 ==========

@app.route('/api/upload/qdata', methods=['POST'])
def upload_qdata():
    """Q-data 엑셀 파일 업로드"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': '엑셀 파일만 업로드 가능합니다.'}), 400
    
    try:
        # 임시 파일 저장
        upload_folder = 'uploads'
        os.makedirs(upload_folder, exist_ok=True)
        file_path = os.path.join(upload_folder, file.filename)
        file.save(file_path)
        
        # 엑셀 읽기
        df = read_qdata_excel(file_path)
        
        # 데이터베이스 저장
        conn = sqlite3.connect('voc_data.db')
        cursor = conn.cursor()
        
        uploaded_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        inserted_count = 0
        duplicate_count = 0
        
        for _, row in df.iterrows():
            # S/N이 없는 경우 건너뛰기
            if pd.isna(row['serial_number']):
                duplicate_count += 1  # NULL 데이터는 중복으로 카운트
                continue
            
            # 공백 제거 (trim)
            serial_number = str(row['serial_number']).strip()
            
            # 빈 문자열 체크
            if not serial_number:
                duplicate_count += 1
                continue
            
            # log_id는 NULL 허용
            log_id = None
            if not pd.isna(row['log_id']):
                log_id = str(row['log_id']).strip() if str(row['log_id']).strip() else None
            
            try:
                cursor.execute('''
                    INSERT INTO q_data (
                        service_date, process_type, repair_name, repair_detail,
                        detail_content, model_name, serial_number, log_id,
                        sw_before, sw_after, uploaded_date
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    row['service_date'],
                    row['process_type'],
                    row['repair_name'],
                    row['repair_detail'],
                    row['detail_content'],
                    row['model_name'],
                    serial_number,  # 정제된 값
                    log_id,         # NULL 허용
                    row['sw_before'],
                    row['sw_after'],
                    uploaded_date
                ))
                inserted_count += 1
            except sqlite3.IntegrityError:
                # 중복 데이터 (S/N 기준)
                duplicate_count += 1
                continue
        
        conn.commit()
        conn.close()
        
        # 임시 파일 삭제
        os.remove(file_path)
        
        return jsonify({
            'success': True,
            'message': f'Q-data 업로드 완료: {inserted_count}건 저장, {duplicate_count}건 중복',
            'inserted': inserted_count,
            'duplicates': duplicate_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'업로드 실패: {str(e)}'
        }), 500

@app.route('/api/statistics/qdata/model', methods=['GET'])
def get_qdata_model_statistics():
    """모델별 Q-data 통계 (처리유형 분포 포함)"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect('voc_data.db')
    cursor = conn.cursor()
    
    # 기본 쿼리
    query = '''
        SELECT 
            model_name,
            COUNT(*) as count
        FROM q_data
        WHERE 1=1
    '''
    params = []
    
    # 날짜 필터
    if start_date:
        query += ' AND service_date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND service_date <= ?'
        params.append(end_date)
    
    query += ' GROUP BY model_name ORDER BY count DESC'
    
    cursor.execute(query, params)
    results = cursor.fetchall()
    
    # 각 모델별 처리유형 분포 조회
    data = []
    for row in results:
        model_name = row[0]
        count = row[1]
        
        # 처리유형별 건수 조회
        process_query = '''
            SELECT process_type, COUNT(*) as type_count
            FROM q_data
            WHERE model_name = ?
        '''
        process_params = [model_name]
        
        if start_date:
            process_query += ' AND service_date >= ?'
            process_params.append(start_date)
        if end_date:
            process_query += ' AND service_date <= ?'
            process_params.append(end_date)
        
        process_query += ' GROUP BY process_type'
        
        cursor.execute(process_query, process_params)
        process_results = cursor.fetchall()
        
        process_types = {}
        for p_row in process_results:
            if p_row[0]:  # process_type이 NULL이 아닌 경우만
                process_types[p_row[0]] = p_row[1]
        
        data.append({
            'model_name': model_name,
            'count': count,
            'process_types': process_types
        })
    
    conn.close()
    
    return jsonify(data)

@app.route('/api/statistics/qdata/monthly', methods=['GET'])
def get_qdata_monthly_statistics():
    """월별 Q-data 전체 건수"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect('voc_data.db')
    cursor = conn.cursor()
    
    query = '''
        SELECT 
            strftime('%Y-%m', service_date) as month,
            COUNT(*) as count
        FROM q_data
        WHERE 1=1
    '''
    params = []
    
    if start_date:
        query += ' AND service_date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND service_date <= ?'
        params.append(end_date)
    
    query += ' GROUP BY month ORDER BY month'
    
    cursor.execute(query, params)
    results = cursor.fetchall()
    
    conn.close()
    
    data = [{'month': row[0], 'count': row[1]} for row in results]
    return jsonify(data)

@app.route('/api/statistics/qdata/models/monthly', methods=['POST'])
def get_qdata_models_monthly():
    """선택된 모델들의 월별 Q-data 건수"""
    data = request.get_json()
    model_names = data.get('model_names', [])
    
    if not model_names:
        return jsonify({})
    
    conn = sqlite3.connect('voc_data.db')
    cursor = conn.cursor()
    
    result = {}
    
    for model_name in model_names:
        query = '''
            SELECT 
                strftime('%Y-%m', service_date) as month,
                COUNT(*) as count
            FROM q_data
            WHERE model_name = ?
            GROUP BY month
            ORDER BY month
        '''
        
        cursor.execute(query, (model_name,))
        rows = cursor.fetchall()
        
        result[model_name] = [{'month': row[0], 'count': row[1]} for row in rows]
    
    conn.close()
    
    return jsonify(result)

@app.route('/api/export/qdata/excel', methods=['GET'])
def export_qdata_excel():
    """Q-data 엑셀 다운로드"""
    model_name = request.args.get('model_name')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect('voc_data.db')
    
    query = 'SELECT * FROM q_data WHERE 1=1'
    params = []
    
    if model_name:
        query += ' AND model_name = ?'
        params.append(model_name)
    if start_date:
        query += ' AND service_date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND service_date <= ?'
        params.append(end_date)
    
    query += ' ORDER BY service_date DESC'
    
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    
    # 엑셀 파일 생성
    output_file = 'qdata_export.xlsx'
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    return send_file(
        output_file,
        as_attachment=True,
        download_name=f'qdata_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/qdata/check-duplicates', methods=['GET'])
def check_qdata_duplicates():
    """Q-data 중복 확인 (serial_number 기준)"""
    conn = sqlite3.connect('voc_data.db')
    cursor = conn.cursor()
    
    # 중복된 S/N 찾기
    cursor.execute('''
        SELECT serial_number, COUNT(*) as count
        FROM q_data
        GROUP BY serial_number
        HAVING count > 1
        ORDER BY count DESC
    ''')
    
    duplicates = cursor.fetchall()
    conn.close()
    
    if duplicates:
        return jsonify({
            'success': True,
            'has_duplicates': True,
            'duplicate_count': len(duplicates),
            'duplicates': [
                {
                    'serial_number': row[0],
                    'count': row[1]
                } for row in duplicates
            ]
        })
    else:
        return jsonify({
            'success': True,
            'has_duplicates': False,
            'duplicate_count': 0
        })

@app.route('/api/qdata/remove-duplicates', methods=['POST'])
def remove_qdata_duplicates():
    """Q-data 중복 제거 (serial_number 기준, 가장 최근 업로드만 유지)"""
    conn = sqlite3.connect('voc_data.db')
    cursor = conn.cursor()
    
    # 중복 제거: S/N이 같은 경우 가장 최근 업로드만 유지
    cursor.execute('''
        DELETE FROM q_data
        WHERE id NOT IN (
            SELECT MAX(id)
            FROM q_data
            GROUP BY serial_number
        )
    ''')
    
    removed_count = cursor.rowcount
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'{removed_count}건의 중복 데이터를 제거했습니다.',
        'removed': removed_count
    })

@app.route('/api/reset/qdata', methods=['POST'])
def reset_qdata_data():
    """Q-data 전체 데이터 초기화"""
    try:
        conn = sqlite3.connect('voc_data.db')
        cursor = conn.cursor()
        
        # Q-data 데이터 삭제 건수 확인
        cursor.execute("SELECT COUNT(*) FROM q_data")
        qdata_count = cursor.fetchone()[0]
        
        # Q-data 전체 삭제
        cursor.execute("DELETE FROM q_data")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Q-data 데이터 초기화 완료: {qdata_count}건 삭제'
        })
    except Exception as e:
        return jsonify({'error': f'Q-data 초기화 실패: {str(e)}'}), 500

# <<<< qdata_backend.py 라우트 끝 >>>>>>>


if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
