"""
Q-data 시스템 추가 기능
app.py에 추가할 코드들입니다.
"""

import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, send_file
import sqlite3
import os

# ========== 유틸리티 함수 ==========

def convert_qdata_date(date_str):
    """
    Q-data 날짜 변환: 260209 → 2026-02-09
    """
    if pd.isna(date_str):
        return None
    try:
        date_str = str(int(date_str))  # 260209
        year = '20' + date_str[:2]     # 2026
        month = date_str[2:4]          # 02
        day = date_str[4:6]            # 09
        return f"{year}-{month}-{day}"
    except:
        return None

def read_qdata_excel(file_path):
    """
    Q-data 엑셀 파일 읽기 (DRM 처리 포함)
    - 9행부터 데이터 시작 (header=8)
    - F, M, P, Q, T, Z, AD, AR, BE, BF 열 읽기
    - DRM 파일 처리 (8가지 방법 fallback)
    """
    # 열 인덱스 (0부터 시작)
    # F=5, M=12, P=15, Q=16, T=19, Z=25, AD=29, AR=43, BE=50, BF=51
    usecols = [5, 12, 15, 16, 19, 25, 29, 43, 50, 51]
    df = None
    
    # DRM 처리 - 8가지 방법 시도
    methods = [
        # Method 1: openpyxl (기본)
        lambda: pd.read_excel(file_path, header=8, usecols=usecols, engine='openpyxl'),
        
        # Method 2: xlrd
        lambda: pd.read_excel(file_path, header=8, usecols=usecols, engine='xlrd'),
        
        # Method 3: pyxlsb (xlsb 형식)
        lambda: pd.read_excel(file_path, header=8, usecols=usecols, engine='pyxlsb'),
        
        # Method 4: 임시 파일로 복사 후 읽기
        lambda: read_via_temp_file(file_path, header=8, usecols=usecols),
        
        # Method 5: xlwings (Windows only)
        lambda: read_via_xlwings(file_path, header=8, usecols=usecols),
        
        # Method 6: win32com (Windows only)
        lambda: read_via_win32com(file_path, header=8, usecols=usecols),
        
        # Method 7: 메모리 스트림
        lambda: read_via_memory_stream(file_path, header=8, usecols=usecols),
        
        # Method 8: openpyxl 직접 사용
        lambda: read_via_openpyxl_direct(file_path, header=8, usecols=usecols)
    ]
    
    last_error = None
    for i, method in enumerate(methods, 1):
        try:
            df = method()
            if df is not None and not df.empty:
                print(f"✓ Q-data DRM 해제 성공 (방법 {i})")
                break
        except Exception as e:
            last_error = e
            continue
    
    if df is None:
        raise Exception(f"Q-data 엑셀 파일 읽기 실패 (DRM): {last_error}")
    
    # 컬럼 이름 매핑 (실제 헤더 이름과 관계없이 순서대로 매핑)
    df.columns = [
        'service_date',    # F열
        'process_type',    # M열
        'repair_name',     # P열
        'repair_detail',   # Q열
        'detail_content',  # T열
        'model_name',      # Z열
        'serial_number',   # AD열
        'log_id',          # AR열
        'sw_before',       # BE열
        'sw_after'         # BF열
    ]
    
    # 날짜 변환
    df['service_date'] = df['service_date'].apply(convert_qdata_date)
    
    # 빈 행 제거
    df = df.dropna(subset=['service_date', 'model_name'])
    
    return df

def read_via_temp_file(file_path, header, usecols):
    """Method 4: 임시 파일로 복사 후 읽기"""
    import tempfile
    import shutil
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp_path = tmp.name
        shutil.copy2(file_path, tmp_path)
        
        try:
            df = pd.read_excel(tmp_path, header=header, usecols=usecols, engine='openpyxl')
            return df
        finally:
            try:
                os.remove(tmp_path)
            except:
                pass

def read_via_xlwings(file_path, header, usecols):
    """Method 5: xlwings 사용 (Windows only)"""
    try:
        import xlwings as xw
        
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        ws = wb.sheets[0]
        
        # 데이터 읽기 (9행부터)
        data = ws.range('A9').expand().value
        
        wb.close()
        app.quit()
        
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # 필요한 열만 선택
        return df.iloc[:, usecols]
    except:
        raise

def read_via_win32com(file_path, header, usecols):
    """Method 6: win32com 사용 (Windows only)"""
    try:
        import win32com.client
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        
        wb = excel.Workbooks.Open(file_path)
        ws = wb.Worksheets(1)
        
        # 데이터 범위 읽기
        used_range = ws.UsedRange
        data = used_range.Value
        
        wb.Close(False)
        excel.Quit()
        
        df = pd.DataFrame(data[header+1:], columns=data[header])
        
        # 필요한 열만 선택
        return df.iloc[:, usecols]
    except:
        raise

def read_via_memory_stream(file_path, header, usecols):
    """Method 7: 메모리 스트림 사용"""
    import io
    
    with open(file_path, 'rb') as f:
        file_content = f.read()
    
    stream = io.BytesIO(file_content)
    return pd.read_excel(stream, header=header, usecols=usecols, engine='openpyxl')

def read_via_openpyxl_direct(file_path, header, usecols):
    """Method 8: openpyxl 직접 사용"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    # 데이터 읽기
    data = []
    for row in ws.iter_rows(min_row=header+2, values_only=True):
        selected_data = [row[i] for i in usecols if i < len(row)]
        if len(selected_data) == len(usecols):
            data.append(selected_data)
    
    wb.close()
    
    column_names = [
        'service_date', 'process_type', 'repair_name', 'repair_detail',
        'detail_content', 'model_name', 'serial_number', 'log_id',
        'sw_before', 'sw_after'
    ]
    
    return pd.DataFrame(data, columns=column_names)

def init_qdata_table():
    """Q-data 테이블 초기화"""
    conn = sqlite3.connect('voc_database.db')
    cursor = conn.cursor()
    
    # SQL 파일 실행
    with open('create_qdata_table.sql', 'r', encoding='utf-8') as f:
        sql_script = f.read()
        cursor.executescript(sql_script)
    
    conn.commit()
    conn.close()

# ========== API 엔드포인트 ==========

@app.route('/api/upload/qdata', methods=['POST'])
def upload_qdata():
    """Q-data 엑셀 파일 업로드"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': '엑셀 파일만 업로드 가능합니다.'}), 400
    
    try:
        # 임시 파일 저장
        upload_folder = 'uploads'
        os.makedirs(upload_folder, exist_ok=True)
        file_path = os.path.join(upload_folder, file.filename)
        file.save(file_path)
        
        # 엑셀 읽기
        df = read_qdata_excel(file_path)
        
        # 데이터베이스 저장
        conn = sqlite3.connect('voc_database.db')
        cursor = conn.cursor()
        
        uploaded_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        inserted_count = 0
        duplicate_count = 0
        
        for _, row in df.iterrows():
            # S/N이 없는 경우 건너뛰기
            if pd.isna(row['serial_number']):
                duplicate_count += 1  # NULL 데이터는 중복으로 카운트
                continue
            
            # 공백 제거 (trim)
            serial_number = str(row['serial_number']).strip()
            
            # 빈 문자열 체크
            if not serial_number:
                duplicate_count += 1
                continue
            
            # log_id는 NULL 허용
            log_id = None
            if not pd.isna(row['log_id']):
                log_id = str(row['log_id']).strip() if str(row['log_id']).strip() else None
            
            try:
                cursor.execute('''
                    INSERT INTO q_data (
                        service_date, process_type, repair_name, repair_detail,
                        detail_content, model_name, serial_number, log_id,
                        sw_before, sw_after, uploaded_date
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    row['service_date'],
                    row['process_type'],
                    row['repair_name'],
                    row['repair_detail'],
                    row['detail_content'],
                    row['model_name'],
                    serial_number,  # 정제된 값
                    log_id,         # NULL 허용
                    row['sw_before'],
                    row['sw_after'],
                    uploaded_date
                ))
                inserted_count += 1
            except sqlite3.IntegrityError:
                # 중복 데이터 (S/N 기준)
                duplicate_count += 1
                continue
        
        conn.commit()
        conn.close()
        
        # 임시 파일 삭제
        os.remove(file_path)
        
        return jsonify({
            'success': True,
            'message': f'Q-data 업로드 완료: {inserted_count}건 저장, {duplicate_count}건 중복',
            'inserted': inserted_count,
            'duplicates': duplicate_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'업로드 실패: {str(e)}'
        }), 500

@app.route('/api/statistics/qdata/model', methods=['GET'])
def get_qdata_model_statistics():
    """모델별 Q-data 통계 (처리유형 분포 포함)"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect('voc_database.db')
    cursor = conn.cursor()
    
    # 기본 쿼리
    query = '''
        SELECT 
            model_name,
            COUNT(*) as count
        FROM q_data
        WHERE 1=1
    '''
    params = []
    
    # 날짜 필터
    if start_date:
        query += ' AND service_date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND service_date <= ?'
        params.append(end_date)
    
    query += ' GROUP BY model_name ORDER BY count DESC'
    
    cursor.execute(query, params)
    results = cursor.fetchall()
    
    # 각 모델별 처리유형 분포 조회
    data = []
    for row in results:
        model_name = row[0]
        count = row[1]
        
        # 처리유형별 건수 조회
        process_query = '''
            SELECT process_type, COUNT(*) as type_count
            FROM q_data
            WHERE model_name = ?
        '''
        process_params = [model_name]
        
        if start_date:
            process_query += ' AND service_date >= ?'
            process_params.append(start_date)
        if end_date:
            process_query += ' AND service_date <= ?'
            process_params.append(end_date)
        
        process_query += ' GROUP BY process_type'
        
        cursor.execute(process_query, process_params)
        process_results = cursor.fetchall()
        
        process_types = {}
        for p_row in process_results:
            if p_row[0]:  # process_type이 NULL이 아닌 경우만
                process_types[p_row[0]] = p_row[1]
        
        data.append({
            'model_name': model_name,
            'count': count,
            'process_types': process_types
        })
    
    conn.close()
    
    return jsonify(data)

@app.route('/api/statistics/qdata/monthly', methods=['GET'])
def get_qdata_monthly_statistics():
    """월별 Q-data 전체 건수"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect('voc_database.db')
    cursor = conn.cursor()
    
    query = '''
        SELECT 
            strftime('%Y-%m', service_date) as month,
            COUNT(*) as count
        FROM q_data
        WHERE 1=1
    '''
    params = []
    
    if start_date:
        query += ' AND service_date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND service_date <= ?'
        params.append(end_date)
    
    query += ' GROUP BY month ORDER BY month'
    
    cursor.execute(query, params)
    results = cursor.fetchall()
    
    conn.close()
    
    data = [{'month': row[0], 'count': row[1]} for row in results]
    return jsonify(data)

@app.route('/api/statistics/qdata/models/monthly', methods=['POST'])
def get_qdata_models_monthly():
    """선택된 모델들의 월별 Q-data 건수"""
    data = request.get_json()
    model_names = data.get('model_names', [])
    
    if not model_names:
        return jsonify({})
    
    conn = sqlite3.connect('voc_database.db')
    cursor = conn.cursor()
    
    result = {}
    
    for model_name in model_names:
        query = '''
            SELECT 
                strftime('%Y-%m', service_date) as month,
                COUNT(*) as count
            FROM q_data
            WHERE model_name = ?
            GROUP BY month
            ORDER BY month
        '''
        
        cursor.execute(query, (model_name,))
        rows = cursor.fetchall()
        
        result[model_name] = [{'month': row[0], 'count': row[1]} for row in rows]
    
    conn.close()
    
    return jsonify(result)

@app.route('/api/export/qdata/excel', methods=['GET'])
def export_qdata_excel():
    """Q-data 엑셀 다운로드"""
    model_name = request.args.get('model_name')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect('voc_database.db')
    
    query = 'SELECT * FROM q_data WHERE 1=1'
    params = []
    
    if model_name:
        query += ' AND model_name = ?'
        params.append(model_name)
    if start_date:
        query += ' AND service_date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND service_date <= ?'
        params.append(end_date)
    
    query += ' ORDER BY service_date DESC'
    
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    
    # 엑셀 파일 생성
    output_file = 'qdata_export.xlsx'
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    return send_file(
        output_file,
        as_attachment=True,
        download_name=f'qdata_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/qdata/check-duplicates', methods=['GET'])
def check_qdata_duplicates():
    """Q-data 중복 확인 (serial_number 기준)"""
    conn = sqlite3.connect('voc_database.db')
    cursor = conn.cursor()
    
    # 중복된 S/N 찾기
    cursor.execute('''
        SELECT serial_number, COUNT(*) as count
        FROM q_data
        GROUP BY serial_number
        HAVING count > 1
        ORDER BY count DESC
    ''')
    
    duplicates = cursor.fetchall()
    conn.close()
    
    if duplicates:
        return jsonify({
            'success': True,
            'has_duplicates': True,
            'duplicate_count': len(duplicates),
            'duplicates': [
                {
                    'serial_number': row[0],
                    'count': row[1]
                } for row in duplicates
            ]
        })
    else:
        return jsonify({
            'success': True,
            'has_duplicates': False,
            'duplicate_count': 0
        })

@app.route('/api/qdata/remove-duplicates', methods=['POST'])
def remove_qdata_duplicates():
    """Q-data 중복 제거 (serial_number 기준, 가장 최근 업로드만 유지)"""
    conn = sqlite3.connect('voc_database.db')
    cursor = conn.cursor()
    
    # 중복 제거: S/N이 같은 경우 가장 최근 업로드만 유지
    cursor.execute('''
        DELETE FROM q_data
        WHERE id NOT IN (
            SELECT MAX(id)
            FROM q_data
            GROUP BY serial_number
        )
    ''')
    
    removed_count = cursor.rowcount
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'{removed_count}건의 중복 데이터를 제거했습니다.',
        'removed': removed_count
    })

# ========== 초기화 ==========
# 앱 시작 시 Q-data 테이블 생성
# init_qdata_table()  # app.py의 적절한 위치에 추가
